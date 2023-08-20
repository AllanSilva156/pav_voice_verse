#!/usr/bin/env python3

import openpyxl
import os
import re
import sys

import subprocess
import pandas as pd
from pydub import AudioSegment
from datetime import timedelta
from moviepy.editor import *
import numpy as np
from scipy.signal import butter, lfilter
from pydub import AudioSegment
import moviepy.editor as mp
# pip install openpyxl
# pip install --upgrade pandas
# pip install --upgrade pydub
# pip install --upgrade moviepy
# sudo apt install ffmpeg
# pip install --upgrade scipy
# pip install numpy
# pip install scipy

def is_time_format(value):
    try:
        time_parts = value.split(':')
        if len(time_parts) != 4:
            return False
        int(time_parts[0]), int(time_parts[1]), int(time_parts[2]), int(time_parts[3])
        return True
    except ValueError:
        return False

def str_to_timedelta(timecode_str: str, frame_rate: float = 30.0) -> timedelta:
    hours, minutes, seconds, frames = map(int, timecode_str.split(':'))
    total_seconds = hours * 3600 + minutes * 60 + seconds + frames / frame_rate
    return timedelta(seconds=total_seconds)

def timecode_to_timedelta(timecode_str, frame_rate: float = 30.0):
    hours, minutes, seconds, frames = timecode_str.split(':')
    return timedelta(hours=int(hours), minutes=int(minutes), seconds=int(seconds), milliseconds=int(frames) * (1000 / frame_rate))


def read_xlsx_file(file_name):
    """
    Read an xlsx file and parse its content into a list of lists.

    Args:
        file_name (str): The name of the xlsx file.

    Returns:
        table_data (list): A list of lists with the parsed content.
    """
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    max_row = sheet.max_row
    
    # Timecode pattern
    timecode_pattern = re.compile(r'^\d{2}:\d{2}:\d{2}(:\d{2})?$')

    # Parse the table data
    table_data = []
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, 3):  # Only read the first two columns
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None:
                cell_value = str(cell_value).replace('\n', ' ').strip()

                # Check if the first column is a timecode
                if col == 1:
                    if not timecode_pattern.match(cell_value):
                        raise ValueError(f"Invalid timecode format in Excel row {row}, column {col}: {cell_value}")

                # Check if the second column is a timecode
                if col == 2:
                    if timecode_pattern.match(cell_value):
                        raise ValueError(f"Invalid text format in Excel row {row}, column {col}: {cell_value}")

                row_data.append(cell_value)
        if row_data:
            table_data.append(row_data)

    return table_data


def convert_mp4_to_wav(mp4_path, wav_path):
    """
    Args:
    mp4_path (str): The path to the input mp4 video file.
    wav_path (str): The path where the output wav audio file will be saved.

    Returns:
    bool: True if the conversion was successful, False otherwise.
    """
    try:
        audio_clip = AudioFileClip(mp4_path)
        audio_clip.write_audiofile(wav_path)
        print("extracting wave from original videoclip")
        return True
    except Exception as e:
        print(f"Error: {e}")
        return False
    
    
def load_audio(audio_file):
    """Load an audio file into a Pydub AudioSegment object."""
    return AudioSegment.from_file(audio_file)


def apply_gain_to_clip(audio_clip, target_dBFS=-3.0):
    """
    Apply gain to an audio clip to match the target dBFS.

    Args:
        audio_clip (pydub.AudioSegment): The audio clip to apply gain to.
        target_dBFS (float): The target dBFS for the audio clip. Defaults to -3.0.

    Returns:
        pydub.AudioSegment: The audio clip with the applied gain.
    """
    dBFS_difference = target_dBFS - audio_clip.max_dBFS
    return audio_clip.apply_gain(dBFS_difference)


def apply_speedup_to_clip(audio_clip, text, rapidospeed, muitorapidospeed):
    """
    Apply speedup to an audio clip based on the text content.

    Args:
        audio_clip (pydub.AudioSegment): The audio clip to apply speedup to.
        text (str): The text containing speedup information.
        rapidospeed (float): The speedup factor for "(Rápido)".
        muitorapidospeed (float): The speedup factor for "(Muito Rápido)".

    Returns:
        pydub.AudioSegment: The audio clip with the applied speedup.
    """
    text = text.lower()
    if 'muito rápido' in text:
        return audio_clip.speedup(muitorapidospeed)
    elif 'rápido' in text:
        return audio_clip.speedup(rapidospeed)
    else:
        return audio_clip
    
def timecode_to_formatted_string(timecode, frame_rate=30):
    """
    Convert a timedelta timecode to a formatted string with the format hh:mm:ss:frames.

    Args:
        timecode (datetime.timedelta): The timedelta object representing the timecode.
        frame_rate (int): The frame rate to calculate the number of frames. Defaults to 30.

    Returns:
        str: The formatted timecode string.
    """
    timecode = str_to_timedelta(timecode)
    total_seconds = int(timecode.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    frames = int((timecode - timedelta(seconds=total_seconds)).total_seconds() * frame_rate)
    formatted_timecode = f"{hours:02d}:{minutes:02d}:{seconds:02d}:{frames:02d}"

    return formatted_timecode

def export_mixed_audio(mixed_audio, output_file):
    """Export a mixed audio track to a wav file."""
    print("exporting mixed audio")
    mixed_audio.export(output_file, format="wav")

def mp4_replace_audio(mp4_path, wav_path, output_path):
    """Replace the audio track of an MP4 video file with a new audio file in WAV format, without re-encoding the video.

    Args:
        mp4_path (str): The path to the input MP4 video file.
        wav_path (str): The path to the input WAV audio file.
        output_path (str): The path to the output MP4 video file with the new audio track.

    Returns:
        bool: True if the audio was replaced successfully, False otherwise.
    """
    try:
        print("substituting audio on the final videoclip without reencoding video")
        result = subprocess.run(['ffmpeg', '-i', mp4_path, '-i', wav_path, '-c:v', 'copy', '-c:a', 'aac', '-map', '0:v', '-map', '1:a', '-shortest', '-y', output_path], capture_output=True)
        if result.returncode != 0:
            print(f"Error: {result.stderr.decode('utf-8')}")
            return False
        else:
            print("Audio replaced successfully.")
            return True
    except Exception as e:
        print(f"Error: {e}")
        return False
    


def filter_band_stop(audio_segment):
    # Convert the AudioSegment object to a NumPy array
    audio_data = np.array(audio_segment.get_array_of_samples(), dtype=np.float32)
    sample_rate = audio_segment.frame_rate

    # Design a Butterworth bandstop filter
    nyquist_rate = 0.5 * sample_rate
    low_cutoff = 1000 / nyquist_rate
    high_cutoff = 2000 / nyquist_rate
    order = 1
    b, a = butter(order, [low_cutoff, high_cutoff], btype='bandstop')

    # Apply the filter to the audio data
    filtered_audio_data = lfilter(b, a, audio_data)

    # Convert the filtered NumPy array back to an AudioSegment object
    filtered_audio_segment = AudioSegment(
        (filtered_audio_data * np.iinfo(np.int16).max).astype(np.int16).tobytes(),
        frame_rate=sample_rate,
        sample_width=audio_segment.sample_width,
        channels=audio_segment.channels
    )

    return filtered_audio_segment

def equalize_audio(audio_clip, equalization_profile):
    equalized_audio = audio_clip

    for low_freq, high_freq, gain_db in equalization_profile:
        # Apply the filter_white_noise function to the audio_clip
        band_audio = filter_band_stop(audio_clip)

        # Apply the gain adjustment
        band_audio = band_audio + gain_db

    return equalized_audio

def load_script(script_file_path):
    script_data = pd.read_excel(script_file_path)
    timecodes = script_data.iloc[:, 0].values
    texts = script_data.iloc[:, 1].values
    script = list(zip(timecodes, texts))
    return script

def apply_preprocessing(audio_clip):
    audio_clip = apply_gain_to_clip(audio_clip)
    audio_clip = apply_speedup_to_clip(audio_clip)
    audio_clip = equalize_audio(audio_clip)
    # Add any other preprocessing functions as needed
    return audio_clip

def mix_audio(synthesized_audio, base_audio):
    mixed_audio = synthesized_audio.overlay(base_audio)
    return mixed_audio

def main():
    # Step 1: Load the script
    script_file_path = 'path/to/your/script.xlsx'
    script_data = load_script(script_file_path)

    # Step 2: Load the synthesized audio
    synthesized_audio_file = 'path/to/your/synthesized_audio.wav'
    synthesized_audio = AudioSegment.from_wav(synthesized_audio_file)

    # Step 3: Load the video
    video_file = 'path/to/your/video.mp4'
    video_clip = mp.VideoFileClip(video_file)

    # Step 4: Apply preprocessing operations
    synthesized_audio = apply_preprocessing(synthesized_audio)

    # Step 5: Mix the synthesized audio with the base audio
    base_audio = video_clip.audio
    mixed_audio = mix_audio(synthesized_audio, base_audio)

    # Step 6: Replace the audio track of the original MP4
    new_video_clip = video_clip.set_audio(mixed_audio)

    # Step 7: Save the result as a new MP4 file
    output_file = 'path/to/your/output.mp4'
    new_video_clip.write_videofile(output_file)

    print('Mixing completed successfully!')
if __name__ == '__main__':
    main()


"""esse script é baseado no desenvolvido por Alvaro Antelo em 17 Mar 2023 pelo MediaTechLAB da Globo TV"""